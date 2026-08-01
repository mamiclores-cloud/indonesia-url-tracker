"""Google Sheet pull/mirror and the relocation-based write-back executor.

The xlsx reader and the Sheets API reader both produce the same intermediate
row shape, so parse_rows()/sync_mirror() are shared between offline tests and
the live sheet:  {"row": int, "vals": {"A".."K": str}, "h_url": str|None,
"e_num": float|None}
"""
import json
import logging
import os
import re
import threading

from . import config as cfg
from . import db
from . import linkparse

log = logging.getLogger(__name__)

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
# L 欄 = 搜尋關鍵字驗證欄（客戶 1.2；J/K 是人工欄位不可用）。read-only 解析
# 仍只看 A-I；L 只由程式寫入。
COLS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

# Machine-written note vocabulary. I-column values outside this set are human
# notes and must never be overwritten (SPEC: 防蓋規則).
# 舊詞彙（invalid/unlisted/high cost）必須保留——表上既有的舊機器 note 才能
# 被新規則覆寫，否則會被當成人工備註永遠洗不掉。
MACHINE_NOTES = {"invalid", "sold out", "unlisted", "high cost",
                 "high_cost", "link error", "new"}


class GoogleAuthNeeded(Exception):
    pass


# ---------------------------------------------------------------- auth ------

_auth_lock = threading.Lock()
_auth_flow_state = {"running": False, "error": None}


def get_creds():
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request

    if not os.path.exists(cfg.GOOGLE_TOKEN_PATH):
        raise GoogleAuthNeeded("尚未完成 Google 授權")
    creds = Credentials.from_authorized_user_file(cfg.GOOGLE_TOKEN_PATH, SCOPES)
    if not creds.valid:
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(cfg.GOOGLE_TOKEN_PATH, "w", encoding="utf-8") as f:
                f.write(creds.to_json())
        else:
            raise GoogleAuthNeeded("Google 授權已失效，請重新授權")
    return creds


def auth_status():
    try:
        get_creds()
        return {"ok": True, "flow_running": _auth_flow_state["running"],
                "error": _auth_flow_state["error"]}
    except GoogleAuthNeeded as e:
        return {"ok": False, "reason": str(e),
                "has_client_secret": os.path.exists(cfg.CLIENT_SECRET_PATH),
                "flow_running": _auth_flow_state["running"],
                "error": _auth_flow_state["error"]}


def start_auth_flow():
    """Run the InstalledAppFlow in a background thread (opens a browser)."""
    if not os.path.exists(cfg.CLIENT_SECRET_PATH):
        raise GoogleAuthNeeded(f"找不到 {cfg.CLIENT_SECRET_PATH}，請先放入 OAuth 用戶端憑證")
    with _auth_lock:
        if _auth_flow_state["running"]:
            return
        _auth_flow_state.update(running=True, error=None)

    def run():
        try:
            from google_auth_oauthlib.flow import InstalledAppFlow
            flow = InstalledAppFlow.from_client_secrets_file(cfg.CLIENT_SECRET_PATH, SCOPES)
            creds = flow.run_local_server(port=0, open_browser=True,
                                          authorization_prompt_message="")
            with open(cfg.GOOGLE_TOKEN_PATH, "w", encoding="utf-8") as f:
                f.write(creds.to_json())
            log.info("Google OAuth completed")
        except Exception as e:  # surfaced in settings UI
            log.exception("Google OAuth flow failed")
            _auth_flow_state["error"] = str(e)
        finally:
            _auth_flow_state["running"] = False

    threading.Thread(target=run, name="google-oauth", daemon=True).start()


def _session():
    from google.auth.transport.requests import AuthorizedSession
    return AuthorizedSession(get_creds())


# ---------------------------------------------------------------- readers ---

def read_xlsx(path=None):
    """Offline reader for the downloaded copy — same shape as read_google()."""
    import openpyxl
    path = path or cfg.XLSX_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[cfg.get("worksheet_name")]
    rows = []
    for r in range(1, ws.max_row + 1):
        vals, h_url, e_num = {}, None, None
        for i, col in enumerate(COLS, start=1):
            cell = ws.cell(row=r, column=i)
            v = cell.value
            if col == "E" and isinstance(v, (int, float)):
                e_num = float(v)
            if col == "H" and cell.hyperlink is not None:
                h_url = cell.hyperlink.target
            vals[col] = "" if v is None else str(v).strip()
        rows.append({"row": r, "vals": vals, "h_url": h_url, "e_num": e_num})
    return rows


FIELDS_MASK = ("sheets(properties(sheetId,title),"
               "data(startRow,startColumn,rowData(values("
               "formattedValue,hyperlink,"
               "effectiveValue(numberValue,stringValue),"
               "userEnteredValue(formulaValue),"
               "textFormatRuns(format(link(uri)))))))")


def read_google(ranges=None):
    """Pull grid data (values + hyperlinks) from the live sheet."""
    ws = cfg.get("worksheet_name")
    ranges = ranges or [f"{ws}!A1:L"]
    sess = _session()
    params = [("includeGridData", "true"), ("fields", FIELDS_MASK)]
    params += [("ranges", r) for r in ranges]
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{cfg.get('sheet_id')}"
    resp = sess.get(url, params=params, timeout=180)
    resp.raise_for_status()
    payload = resp.json()
    sheet = None
    for s in payload.get("sheets", []):
        if s.get("properties", {}).get("title") == ws:
            sheet = s
            break
    if sheet is None:
        raise RuntimeError(f"找不到分頁「{ws}」")
    db.setting_set("worksheet_sheet_id", sheet["properties"]["sheetId"])

    rows = []
    for block in sheet.get("data", []):
        start_row = block.get("startRow", 0)
        start_col = block.get("startColumn", 0)
        for i, rd in enumerate(block.get("rowData", [])):
            r = start_row + i + 1
            vals, h_url, e_num = {c: "" for c in COLS}, None, None
            for j, cv in enumerate(rd.get("values", [])):
                ci = start_col + j
                if ci >= len(COLS):
                    break
                col = COLS[ci]
                text = (cv.get("formattedValue") or "").strip()
                vals[col] = text
                eff = cv.get("effectiveValue", {})
                if col == "E" and "numberValue" in eff:
                    e_num = eff["numberValue"]
                if col == "H":
                    h_url = cv.get("hyperlink")
                    if not h_url:
                        for run in cv.get("textFormatRuns", []) or []:
                            uri = run.get("format", {}).get("link", {}).get("uri")
                            if uri:
                                h_url = uri
                                break
                    if not h_url:
                        formula = cv.get("userEnteredValue", {}).get("formulaValue", "")
                        m = re.search(r'HYPERLINK\s*\(\s*"([^"]+)"', formula, re.I)
                        if m:
                            h_url = m.group(1)
                    if not h_url and text.startswith("http"):
                        h_url = text
            rows.append({"row": r, "vals": vals, "h_url": h_url, "e_num": e_num})
    rows.sort(key=lambda x: x["row"])
    return rows


def get_sheet_id_num():
    v = db.setting_get("worksheet_sheet_id")
    if v is not None:
        return int(v)
    read_google(ranges=[f"{cfg.get('worksheet_name')}!A1:A1"])
    return int(db.setting_get("worksheet_sheet_id"))


# ---------------------------------------------------------------- parsing ---

def parse_price(row):
    if row["e_num"] is not None:
        return int(round(row["e_num"]))
    s = row["vals"].get("E", "")
    if not s:
        return None
    digits = re.sub(r"[^\d]", "", s)
    return int(digits) if digits else None


def is_link_row(row):
    if row["h_url"]:
        return True
    return row["vals"].get("H", "").startswith("http")


def parse_rows(rows):
    """Shared parser. Returns (products, links, last_row_by_code).

    products: {code: {"item_name": str, "first_row": int}}
    links: list of dicts ready for sync_mirror()
    last_row_by_code: 1-based last sheet row belonging to each code's group
    (used by the insert executor when computed from a *fresh* pull).
    """
    products, links, last_row = {}, [], {}
    code = None
    prev_code = None
    for row in rows:
        if row["row"] == 1:  # header
            continue
        a = row["vals"].get("A", "").strip()
        if a:
            code = a
        if not code:
            prev_code = None
            continue
        b = row["vals"].get("B", "").strip()
        if code not in products:
            products[code] = {"item_name": b, "first_row": row["row"]}
        elif b and not products[code]["item_name"]:
            products[code]["item_name"] = b
        # Insert point = end of the code's FIRST contiguous block only. A code
        # that reappears far down the sheet (a duplicate/second block) must not
        # drag new-link inserts away from its main group — otherwise the row is
        # written 500 rows away from where the user sees the product.
        has_content = a or any(row["vals"].get(c, "") for c in ("B", "C", "D", "E", "G", "H", "I"))
        if has_content:
            if code not in last_row or prev_code == code:
                last_row[code] = row["row"]
            prev_code = code
        if not is_link_row(row):
            continue
        # 0801 修正單：儲存格顯示文字優先（客戶要求抓儲存格中的連結文字，
        # 不抓附掛的短網址 hyperlink）；文字不是 URL 才退回 hyperlink/公式。
        h_text = row["vals"]["H"].strip()
        raw_url = h_text if h_text.startswith("http") else (row["h_url"] or h_text)
        links.append({
            "product_code": code,
            "sheet_row": row["row"],
            "raw_url": raw_url,
            "page_name": row["vals"].get("C", ""),
            "variant_text": row["vals"].get("D", ""),
            "price_idr": parse_price(row),
            "supplier": row["vals"].get("G", ""),
            "note": row["vals"].get("I", ""),
        })
    return products, links, last_row


def enrich_link_ids(link):
    """Static (no-network) id extraction; short links use url_cache only."""
    url = link["raw_url"]
    final = url
    if not linkparse.is_shopee(url):
        row = db.q1("SELECT final_url FROM url_cache WHERE short_url=?", (url,))
        final = row["final_url"] if row and row["final_url"] else None
    ids = linkparse.parse_shopee_url(final) if final else None
    if ids:
        link.update(shopid=ids["shopid"], itemid=ids["itemid"], model_id=ids["model_id"],
                    dedupe_key=linkparse.dedupe_key(ids["shopid"], ids["itemid"]),
                    canonical_url=linkparse.canonical_url(ids["shopid"], ids["itemid"], ids["model_id"]))
    else:
        link.update(shopid=None, itemid=None, model_id=None, dedupe_key="", canonical_url=None)
    return link


def seed_baselines(parse_min=None):
    """Seed products.baseline_idr for products that don't have one yet.

    客戶規則：基準價 = 該商品「未執行查詢前」表上 E 欄最低價，只定植一次，
    之後僅人工可改。程式已回寫過部分 E 欄，所以優先讀 xlsx 原始快照；
    快照沒有的商品用 parse_min（本次拉取各 code 最低價）或 mirror 最低價
    （排除程式插入的 origin='found' 列）。
    """
    unseeded = [r["code"] for r in db.q("SELECT code FROM products WHERE baseline_idr IS NULL")]
    if not unseeded:
        return {"seeded": 0}
    snapshot_min = {}
    if os.path.exists(cfg.XLSX_PATH):
        try:
            _, snap_links, _ = parse_rows(read_xlsx())
            for lk in snap_links:
                p = lk["price_idr"]
                if p and p > 0:
                    code = lk["product_code"]
                    if code not in snapshot_min or p < snapshot_min[code]:
                        snapshot_min[code] = p
        except Exception:
            log.exception("xlsx snapshot read failed; baseline falls back to mirror prices")
    seeded = 0
    for code in unseeded:
        base = snapshot_min.get(code) or (parse_min or {}).get(code)
        if base is None:
            row = db.q1("SELECT MIN(price_idr) AS m FROM links WHERE product_code=? AND active=1 "
                        "AND origin='sheet' AND price_idr>0", (code,))
            base = row["m"] if row else None
        if base:
            db.x("UPDATE products SET baseline_idr=? WHERE code=? AND baseline_idr IS NULL",
                 (base, code))
            seeded += 1
    if seeded:
        log.info("baseline seeded for %d products", seeded)
    return {"seeded": seeded}


def sync_mirror(rows):
    """Upsert parsed sheet rows into SQLite.

    身分比對（0801 修正單）：先 (product_code, raw_url) 精確比對；沒中再用
    dedupe_key（shopid.itemid）在同商品內找未匹配的舊列「就地更新 raw_url」——
    同一儲存格在不同次拉取抽出的 URL 字串可能不同（hyperlink vs 文字、編碼
    差異），不能因此把舊列當成已從表上刪除。最後做幻影收斂：同 dedupe_key
    的殘餘 inactive 舊列併入本次匹配到的列（純 DB 整理，Sheet 永遠不動）。
    """
    products, links, _ = parse_rows(rows)
    c = db.conn()
    with c:
        for code, meta in products.items():
            c.execute(
                "INSERT INTO products(code, item_name, updated_at) VALUES(?,?,?) "
                "ON CONFLICT(code) DO UPDATE SET item_name=excluded.item_name, updated_at=excluded.updated_at",
                (code, meta["item_name"], db.now()))
        # Deactivate ALL links first, then reactivate whatever the sheet still
        # has. A 'found' link that was written to the sheet but later deleted by
        # a human must also drop to active=0 — otherwise it lingers as a phantom
        # valid link and can be re-picked as a finder source. 本交易內
        # active=1 == 本次拉取有匹配到，dedupe fallback 據此只認領 active=0 列。
        c.execute("UPDATE links SET active=0")
        for lk in links:
            enrich_link_ids(lk)
            row = c.execute("SELECT id FROM links WHERE product_code=? AND raw_url=?",
                            (lk["product_code"], lk["raw_url"])).fetchone()
            if row is None and lk["dedupe_key"]:
                # 同商品同 dedupe_key 且未被本次認領的舊列 = 同一連結換了字串
                # 表示 → 就地更新 raw_url（優先取有檢查歷史者、再最舊）。
                # 精確比對沒中代表沒有列持有新 raw_url，這裡不會撞 UNIQUE。
                row = c.execute(
                    "SELECT id FROM links WHERE product_code=? AND dedupe_key=? AND active=0 "
                    "ORDER BY (last_checked_at IS NULL), id LIMIT 1",
                    (lk["product_code"], lk["dedupe_key"])).fetchone()
                if row is not None:
                    c.execute("UPDATE links SET raw_url=? WHERE id=?", (lk["raw_url"], row["id"]))
            if row is not None:
                c.execute(
                    """UPDATE links SET sheet_row_hint=?,
                         canonical_url=COALESCE(?, canonical_url),
                         shopid=COALESCE(?, shopid), itemid=COALESCE(?, itemid),
                         model_id=COALESCE(?, model_id),
                         dedupe_key=CASE WHEN ?!='' THEN ? ELSE dedupe_key END,
                         page_name=?, variant_text=?, price_idr=?, supplier=?,
                         note=?, active=1, origin='sheet' WHERE id=?""",
                    (lk["sheet_row"], lk["canonical_url"], lk["shopid"], lk["itemid"],
                     lk["model_id"], lk["dedupe_key"], lk["dedupe_key"],
                     lk["page_name"], lk["variant_text"], lk["price_idr"],
                     lk["supplier"], lk["note"], row["id"]))
            else:
                c.execute(
                    """INSERT INTO links(product_code, sheet_row_hint, raw_url, canonical_url,
                         shopid, itemid, model_id, dedupe_key, page_name, variant_text,
                         price_idr, supplier, note, active, origin)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,1,'sheet')""",
                    (lk["product_code"], lk["sheet_row"], lk["raw_url"], lk["canonical_url"],
                     lk["shopid"], lk["itemid"], lk["model_id"], lk["dedupe_key"],
                     lk["page_name"], lk["variant_text"], lk["price_idr"], lk["supplier"], lk["note"]))
        # 幻影收斂：同 (product_code, dedupe_key) 既有本次匹配到的列（active=1）
        # 又有殘餘 inactive 舊列時，舊列只是同一連結的舊字串身分——把檢查歷史
        # 併入倖存者後刪除 DB 列（Sheet 不動）。兩列都在表上（都 active=1）
        # 絕不合併（同商品確實可能有兩列同款連結）；dedupe_key 空字串絕不合併。
        dups = c.execute(
            """SELECT product_code, dedupe_key FROM links WHERE dedupe_key!=''
               GROUP BY product_code, dedupe_key
               HAVING SUM(active)>=1 AND SUM(1-active)>=1""").fetchall()
        for d in dups:
            survivor = c.execute(
                "SELECT id, status FROM links WHERE product_code=? AND dedupe_key=? AND active=1 "
                "ORDER BY (last_checked_at IS NULL), id LIMIT 1",
                (d["product_code"], d["dedupe_key"])).fetchone()
            phantoms = c.execute(
                "SELECT * FROM links WHERE product_code=? AND dedupe_key=? AND active=0",
                (d["product_code"], d["dedupe_key"])).fetchall()
            for ph in phantoms:
                cur = c.execute("SELECT status FROM links WHERE id=?", (survivor["id"],)).fetchone()
                if cur["status"] == "unchecked" and ph["last_checked_at"]:
                    # 倖存者是改名後新建的空殼列時，把幻影的檢查歷史搬過來
                    c.execute(
                        """UPDATE links SET status=?, status_detail=?, last_price_idr=?,
                             prev_price_idr=?, last_checked_at=?, shop_location=?, sold=?,
                             search_keyword=COALESCE(search_keyword, ?) WHERE id=?""",
                        (ph["status"], ph["status_detail"], ph["last_price_idr"],
                         ph["prev_price_idr"], ph["last_checked_at"], ph["shop_location"],
                         ph["sold"], ph["search_keyword"], survivor["id"]))
                c.execute("DELETE FROM links WHERE id=?", (ph["id"],))
    # 新商品第一次進系統時定植基準價（此時表上價格必然是「查詢前」的）
    parse_min = {}
    for lk in links:
        p = lk["price_idr"]
        if p and p > 0:
            code = lk["product_code"]
            if code not in parse_min or p < parse_min[code]:
                parse_min[code] = p
    seed_baselines(parse_min)
    db.setting_set("last_sync_at", db.now())
    return {"products": len(products), "links": len(links)}


def pull_and_sync():
    return sync_mirror(read_google())


def import_xlsx(path=None):
    return sync_mirror(read_xlsx(path))


# ------------------------------------------------------------- write-back ---

def is_machine_note(note: str) -> bool:
    return (note or "").strip().lower() in MACHINE_NOTES


def _col_range(ws, col, row):
    return f"{ws}!{col}{row}"


def _insert_value_batch(ws, new_row, code, p):
    """新插入列的 values:batchUpdate payload（純函式，供測試斷言）。
    0801 修正單：I 欄不再寫「new」——valid 連結 Note 必須空白，插列後 I 本來就空。"""
    h_formula = f'=HYPERLINK("{p["url"]}")'
    batch = [
        # A 欄填上同樣的 product code（使用者要求：找到的新列也要標明所屬商品）
        {"range": f"{ws}!A{new_row}",
         "values": [[p.get("product_code", code)]]},
        {"range": f"{ws}!C{new_row}:E{new_row}",
         "values": [[p.get("page_name", ""), p.get("variant_text", ""), p.get("price_idr")]]},
        {"range": f"{ws}!G{new_row}:H{new_row}",
         "values": [[p.get("supplier", ""), h_formula]]},
    ]
    # 客戶 1.2：把找此連結用的搜尋關鍵字寫在同列 L 欄供人工驗證
    if p.get("search_keyword") and cfg.get("record_keyword_to_sheet", True):
        batch.append({"range": f"{ws}!L{new_row}", "values": [[p["search_keyword"]]]})
    return batch


class Relocator:
    """Fresh-pull row map. Never trusts sheet_row_hint."""

    def __init__(self):
        self.rows = read_google()
        self.products, self.links, self.last_row = parse_rows(self.rows)
        self.by_raw = {}
        self.by_ids = {}
        self.note_at = {}
        for lk in self.links:
            enrich_link_ids(lk)
            self.by_raw.setdefault((lk["product_code"], lk["raw_url"]), []).append(lk["sheet_row"])
            if lk["dedupe_key"]:
                self.by_ids.setdefault(
                    (lk["product_code"], lk["dedupe_key"], lk["model_id"]), []).append(lk["sheet_row"])
            self.note_at[lk["sheet_row"]] = lk["note"]

    def locate(self, product_code, raw_url, dedupe_key=None, model_id=None):
        hits = self.by_raw.get((product_code, raw_url), [])
        if len(hits) == 1:
            return hits[0]
        if dedupe_key:
            hits = self.by_ids.get((product_code, dedupe_key, model_id), [])
            if len(hits) == 1:
                return hits[0]
            loose = [r for (c, d, m), rs in self.by_ids.items()
                     if c == product_code and d == dedupe_key for r in rs]
            if len(loose) == 1:
                return loose[0]
        return None

    def shift_after(self, row_index):
        for m in (self.by_raw, self.by_ids):
            for k, rs in m.items():
                m[k] = [r + 1 if r > row_index else r for r in rs]
        self.note_at = {(r + 1 if r > row_index else r): v for r, v in self.note_at.items()}
        self.last_row = {k: (v + 1 if v > row_index else v) for k, v in self.last_row.items()}


def apply_writes(write_ids, progress=None):
    """Execute approved pending_writes against the live sheet.

    Cell updates (price/note) first in one batch, then row inserts one at a
    time with read-back verification.
    """
    ws = cfg.get("worksheet_name")
    sid = cfg.get("sheet_id")
    sheet_num = get_sheet_id_num()
    sess = _session()
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{sid}"

    writes = [dict(r) for r in db.q(
        f"SELECT * FROM pending_writes WHERE id IN ({','.join('?'*len(write_ids))})", write_ids)]
    reloc = Relocator()
    results = {"written": 0, "failed": 0}

    def fail(w, msg):
        db.x("UPDATE pending_writes SET state='failed', error=? WHERE id=?", (msg, w["id"]))
        results["failed"] += 1
        log.warning("write %s failed: %s", w["id"], msg)

    def mark_written(w):
        db.x("UPDATE pending_writes SET state='written', error=NULL, written_at=? WHERE id=?",
             (db.now(), w["id"]))
        results["written"] += 1

    def mirror_cell_write(w):
        # keep the local mirror in step so the UI reflects the sheet immediately
        # instead of waiting for the next pull
        p = db.jload(w["payload_json"], {})
        raw = p.get("raw_url", "")
        if w["kind"] == "update_price":
            db.x("UPDATE links SET price_idr=? WHERE product_code=? AND raw_url=?",
                 (p.get("price_idr"), w["product_code"], raw))
        else:
            note = "" if w["kind"] == "clear_note" else p.get("note", "")
            db.x("UPDATE links SET note=? WHERE product_code=? AND raw_url=?",
                 (note, w["product_code"], raw))

    # ---- phase 1: cell updates -------------------------------------------
    value_data, done_cell_writes = [], []
    for w in writes:
        if w["kind"] not in ("update_price", "set_note", "clear_note"):
            continue
        p = db.jload(w["payload_json"], {})
        row = reloc.locate(w["product_code"], p.get("raw_url", ""), w["dedupe_key"], w["model_id"])
        if row is None:
            fail(w, "row not found（表格可能已被編輯，請重新拉取後再試）")
            continue
        if w["kind"] == "update_price":
            value_data.append({"range": _col_range(ws, "E", row), "values": [[p["price_idr"]]]})
        else:
            current = reloc.note_at.get(row, "")
            if current and not is_machine_note(current):
                fail(w, f"note conflict：I 欄有人工備註「{current}」，未覆寫")
                continue
            new_note = "" if w["kind"] == "clear_note" else p["note"]
            if (current or "").strip() == new_note.strip():
                mark_written(w)  # already in desired state
                continue
            value_data.append({"range": _col_range(ws, "I", row), "values": [[new_note]]})
        done_cell_writes.append(w)

    if value_data:
        resp = sess.post(f"{base}/values:batchUpdate", json={
            "valueInputOption": "USER_ENTERED", "data": value_data}, timeout=120)
        if resp.status_code != 200:
            for w in done_cell_writes:
                fail(w, f"batchUpdate HTTP {resp.status_code}: {resp.text[:200]}")
            done_cell_writes = []
        else:
            for w in done_cell_writes:
                mark_written(w)
                mirror_cell_write(w)
    else:
        for w in done_cell_writes:
            mark_written(w)
            mirror_cell_write(w)
    if progress:
        progress(len(done_cell_writes), len(writes))

    # ---- phase 2: row inserts (sequential, verified) ---------------------
    inserts = [w for w in writes if w["kind"] == "insert_link_row"]
    for n, w in enumerate(inserts, 1):
        p = db.jload(w["payload_json"], {})
        code = w["product_code"]
        last = reloc.last_row.get(code)
        if last is None:
            fail(w, f"product group not found for {code}")
            continue
        new_row = last + 1
        try:
            reqs = [{"insertDimension": {
                "range": {"sheetId": sheet_num, "dimension": "ROWS",
                          "startIndex": last, "endIndex": last + 1},
                "inheritFromBefore": True}}]
            r1 = sess.post(f"{base}:batchUpdate", json={"requests": reqs}, timeout=60)
            r1.raise_for_status()

            value_batch = _insert_value_batch(ws, new_row, code, p)
            r2 = sess.post(f"{base}/values:batchUpdate", json={
                "valueInputOption": "USER_ENTERED", "data": value_batch}, timeout=60)
            r2.raise_for_status()

            # F formula copied from the row above (SPEC: 向下複製)
            r3 = sess.post(f"{base}:batchUpdate", json={"requests": [{"copyPaste": {
                "source": {"sheetId": sheet_num, "startRowIndex": last - 1, "endRowIndex": last,
                           "startColumnIndex": 5, "endColumnIndex": 6},
                "destination": {"sheetId": sheet_num, "startRowIndex": last, "endRowIndex": last + 1,
                                "startColumnIndex": 5, "endColumnIndex": 6},
                "pasteType": "PASTE_FORMULA"}}]}, timeout=60)
            r3.raise_for_status()

            # read back verification
            r4 = sess.get(f"{base}/values/{ws}!C{new_row}:H{new_row}",
                          params={"valueRenderOption": "FORMULA"}, timeout=60)
            r4.raise_for_status()
            got = (r4.json().get("values") or [[]])[0]
            got_h = got[5] if len(got) > 5 else ""
            if p["url"].split("?")[0] not in str(got_h):
                fail(w, f"read-back mismatch on row {new_row}: {got_h!r}")
                continue
        except Exception as e:
            fail(w, f"insert error: {e}")
            continue

        reloc.shift_after(last)
        reloc.last_row[code] = new_row
        mark_written(w)
        # mirror into links（地區/銷量/關鍵字帶入，待下次檢查以 PDP 值覆蓋）
        db.x("""INSERT INTO links(product_code, sheet_row_hint, raw_url, canonical_url, shopid, itemid,
                  model_id, dedupe_key, page_name, variant_text, price_idr, supplier, note,
                  status, origin, active, last_checked_at, shop_location, sold, search_keyword)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,'found',1,?,?,?,?)
                ON CONFLICT(product_code, raw_url) DO UPDATE SET active=1, sheet_row_hint=excluded.sheet_row_hint""",
             (code, new_row, p["url"], p["url"], p.get("shopid"), p.get("itemid"),
              w["model_id"], w["dedupe_key"] or "", p.get("page_name", ""), p.get("variant_text", ""),
              p.get("price_idr"), p.get("supplier", ""), "", "valid", db.now(),
              p.get("shop_location", ""), p.get("sold"), p.get("search_keyword", "")))
        if p.get("candidate_id"):
            db.x("UPDATE candidates SET state='written' WHERE id=?", (p["candidate_id"],))
        if progress:
            progress(len(done_cell_writes) + n, len(writes))

    return results
